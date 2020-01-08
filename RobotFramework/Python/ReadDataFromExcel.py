import openpyxl
vk=openpyxl.load_workbook("/home/shivampathak/Documents/TestExcel1.xlsx")

def fetch_max_rows(sheetName):
    sh = vk[sheetName]
    return sh.max_row


def fetch_cell_data(sheetName,row,cell):
    sh = vk[sheetName]
    cell=sh.cell(row,cell)
    return cell.value
print(fetch_cell_data("Sheet1",1,1))
print(fetch_max_rows("Sheet1"))
